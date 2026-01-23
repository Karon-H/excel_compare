import os
import pandas as pd
import openpyxl
import difflib
from openpyxl.utils import range_boundaries
from openpyxl.styles import PatternFill, Font, Alignment

class ExcelDiffer:
    """Excel 比对逻辑处理类"""
    
    @staticmethod
    def load_sheets(filepath):
        """读取 Excel 文件的 Sheet 列表"""
        try:
            # 自动识别引擎
            ext = os.path.splitext(filepath)[1].lower()
            if ext in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
                engine = 'openpyxl'
            elif ext in ['.xls']:
                engine = 'xlrd'
            else:
                engine = None
            
            # 使用 context manager 确保文件关闭
            with pd.ExcelFile(filepath, engine=engine) as xl:
                return xl.sheet_names
        except Exception as e:
            # 降级处理：尝试不指定引擎
            try:
                with pd.ExcelFile(filepath) as xl:
                    return xl.sheet_names
            except:
                raise Exception(f"无法获取 Excel Sheet 列表: {e}")

    @staticmethod
    def compare_all_sheets(file1, file2, key_columns=None, header_row=None, has_header=True, progress_callback=None):
        """
        全表自动比对逻辑：
        1. 自动匹配两个文件中的 Sheet (基于名称相似度)
        2. 对匹配成功的 Sheet 进行比对
        3. 返回比对摘要
        """
        # 1. 获取 Sheet 列表
        sheets1 = ExcelDiffer.load_sheets(file1)
        sheets2 = ExcelDiffer.load_sheets(file2)
        
        # 2. 自动匹配 Sheet
        matched_sheets = [] # List of (sheet1, sheet2)
        unmatched1 = list(sheets1)
        unmatched2 = list(sheets2)
        
        # 优先完全匹配
        for s1 in sheets1[:]:
            if s1 in unmatched2:
                matched_sheets.append((s1, s1))
                unmatched1.remove(s1)
                unmatched2.remove(s1)
        
        # 其次模糊匹配 (忽略大小写和空格)
        for s1 in unmatched1[:]:
            s1_clean = s1.strip().lower()
            for s2 in unmatched2[:]:
                if s2.strip().lower() == s1_clean:
                    matched_sheets.append((s1, s2))
                    unmatched1.remove(s1)
                    unmatched2.remove(s2)
                    break
        
        # 3. 执行比对
        results_summary = []
        total = len(matched_sheets)
        for i, (s1, s2) in enumerate(matched_sheets):
            if progress_callback:
                progress_callback(int((i / total) * 100), f"正在比对 Sheet: {s1}...")
                
            try:
                df1 = ExcelDiffer.read_excel_raw(file1, s1, handle_merged=True, header_row=header_row, has_header=has_header)
                df2 = ExcelDiffer.read_excel_raw(file2, s2, handle_merged=True, header_row=header_row, has_header=has_header)
                
                cols, results = ExcelDiffer.compare_dataframes(df1, df2, key_columns=key_columns)
                
                stats = {'added': 0, 'deleted': 0, 'modified': 0, 'equal': 0}
                for r in results:
                    status = r[2].get('status', 'equal')
                    if status in stats:
                        stats[status] += 1
                
                results_summary.append({
                    'sheet1': s1,
                    'sheet2': s2,
                    'stats': stats,
                    'status': 'success'
                })
            except Exception as e:
                results_summary.append({
                    'sheet1': s1,
                    'sheet2': s2,
                    'error': str(e),
                    'status': 'error'
                })
        
        # 记录未匹配的 Sheet
        for s in unmatched1:
            results_summary.append({'sheet1': s, 'sheet2': None, 'status': 'only_in_file1'})
        for s in unmatched2:
            results_summary.append({'sheet1': None, 'sheet2': s, 'status': 'only_in_file2'})
            
        return results_summary

    @staticmethod
    def read_excel_raw(filepath, sheet_name=None, handle_merged=True, header_row=None, has_header=True):
        """
        极致健壮的 Excel 读取逻辑：
        1. 自动识别引擎，支持多种 Excel 格式
        2. 智能匹配 Sheet 名称（防止空格、大小写导致的失败）
        3. 智能表头识别与合并单元格处理
        - header_row: 指定表头所在行号（从 1 开始）。如果为 None，则自动寻找第一个非空行。
        - has_header: 是否有表头。如果为 False，则自动生成列名。
        """
        # 自动识别引擎
        ext = os.path.splitext(filepath)[1].lower()
        if ext in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            engine = 'openpyxl'
        elif ext in ['.xls']:
            engine = 'xlrd'
        else:
            engine = None
        
        # --- 预处理 Sheet 名称 ---
        # 获取所有实际的 Sheet 名称
        try:
            with pd.ExcelFile(filepath, engine=engine) as xl:
                actual_sheets = xl.sheet_names
            
            # 1. 尝试完全匹配
            target_sheet = sheet_name
            if sheet_name not in actual_sheets:
                # 2. 尝试模糊匹配 (不区分大小写，不区分首尾空格)
                s_name = str(sheet_name).strip().lower()
                for s in actual_sheets:
                    if s.strip().lower() == s_name:
                        target_sheet = s
                        break
        except Exception as e:
            # 如果获取列表失败，只能盲猜
            target_sheet = sheet_name

        df = None
        
        # --- 尝试读取 ---
        try:
            df = pd.read_excel(filepath, sheet_name=target_sheet, header=None, engine=engine)
        except Exception as e:
            # 如果按名称读取失败，记录错误并抛出，不再默认跳回第一个
            raise Exception(f"读取工作表 '{sheet_name}' 失败: {e}")

        if df is None or df.empty:
            return pd.DataFrame()

        # --- 合并单元格处理 (仅针对 openpyxl) ---
        if handle_merged and engine == 'openpyxl':
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                # 使用匹配后的名称
                ws = None
                if target_sheet in wb.sheetnames:
                    ws = wb[target_sheet]
                else:
                    # 再次兜底匹配
                    match = [s for s in wb.sheetnames if s.lower().strip() == str(target_sheet).lower().strip()]
                    ws = wb[match[0]] if match else wb.worksheets[0]
                
                if ws:
                    for merged_range in ws.merged_cells.ranges:
                        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                        top_left_value = ws.cell(row=min_row, column=min_col).value
                        for r in range(min_row-1, max_row):
                            for c in range(min_col-1, max_col):
                                if r < len(df) and c < len(df.columns):
                                    df.iloc[r, c] = top_left_value
                wb.close()
            except Exception as merge_err:
                print(f"合并单元格处理提示: {merge_err}")
        
        # --- 移除全空列 (行先保留，用于准确定位表头行) ---
        df = df.dropna(axis=1, how='all')
        
        if not df.empty:
            # --- 表头识别 ---
            if not has_header:
                # 无表头模式：所有非空行都是数据
                df_data = df.dropna(how='all')
                header_values = [f"列{i+1}" for i in range(len(df.columns))]
            elif header_row is not None:
                # 用户指定了行号（1-based），对应原始索引 header_row-1
                target_idx = int(header_row) - 1
                if target_idx in df.index:
                    header_row_idx = target_idx
                else:
                    # 如果指定行是空行被删了，或者越界，则尝试找第一个非空行
                    header_row_idx = df.index[0] if not df.empty else 0
                
                try:
                    header_values = df.loc[header_row_idx]
                    df_data = df.loc[header_row_idx + 1:].dropna(how='all')
                except:
                    # 兜底逻辑
                    header_values = df.iloc[0]
                    df_data = df.iloc[1:].dropna(how='all')
            else:
                # 自动寻找第一个包含非空值的行
                # 注意此时 df 还没 dropna(how='all')，所以我们需要找第一个非空行
                non_empty_df = df.dropna(how='all')
                header_row_idx = non_empty_df.index[0] if not non_empty_df.empty else 0
                
                try:
                    header_values = df.loc[header_row_idx]
                    df_data = df.loc[header_row_idx + 1:].dropna(how='all')
                except:
                    # 兜底逻辑
                    header_values = df.iloc[0]
                    df_data = df.iloc[1:].dropna(how='all')
            
            # 处理列名
            new_columns = []
            seen = {}
            for i, val in enumerate(header_values):
                col_name = str(val).strip() if pd.notna(val) and str(val).strip() != "" else f"列{i+1}"
                if col_name in seen:
                    seen[col_name] += 1
                    col_name = f"{col_name}_{seen[col_name]}"
                else:
                    seen[col_name] = 0
                new_columns.append(col_name)
            
            df_data.columns = new_columns
            return df_data.reset_index(drop=True)
        
        return df

    @staticmethod
    def get_text_diff_html(old_val, new_val):
        """
        计算两个文本的微观差异并返回 HTML 格式
        使用 difflib 实现字符级对比
        """
        # 处理非字符串类型
        s1 = str(old_val) if pd.notna(old_val) else ""
        s2 = str(new_val) if pd.notna(new_val) else ""
        
        if s1 == s2:
            return s1
            
        s = difflib.SequenceMatcher(None, s1, s2)
        html = []
        for tag, i1, i2, j1, j2 in s.get_opcodes():
            if tag == 'equal':
                html.append(s1[i1:i2])
            elif tag == 'delete':
                text = s1[i1:i2]
                html.append(f'<span style="color: #d73a49; text-decoration: line-through; background-color: #ffeef0;">{text}</span>')
            elif tag == 'insert':
                text = s2[j1:j2]
                html.append(f'<span style="color: #22863a; background-color: #e6ffed; font-weight: bold;">{text}</span>')
            elif tag == 'replace':
                text_old = s1[i1:i2]
                text_new = s2[j1:j2]
                html.append(f'<span style="color: #d73a49; text-decoration: line-through; background-color: #ffeef0;">{text_old}</span>')
                html.append(f'<span style="color: #22863a; background-color: #e6ffed; font-weight: bold;">{text_new}</span>')
        
        return "".join(html)

    @staticmethod
    def compare_dataframes(df1, df2, key_columns=None):
        """
        比对两个 DataFrame。
        - 如果 key_columns 为空，使用序列对齐算法 (difflib.SequenceMatcher)。
        - 如果 key_columns 不为空，使用主键比对算法。
        返回格式: (all_columns, results)
        """
        # 填充缺失值为特定字符串，避免 FutureWarning
        df1 = df1.fillna("").infer_objects(copy=False)
        df2 = df2.fillna("").infer_objects(copy=False)
        
        # 校验主键列是否存在
        if key_columns:
            missing_df1 = [col for col in key_columns if col not in df1.columns]
            missing_df2 = [col for col in key_columns if col not in df2.columns]
            if missing_df1:
                raise Exception(f"旧文件中缺少关键列: {', '.join(missing_df1)}")
            if missing_df2:
                raise Exception(f"新文件中缺少关键列: {', '.join(missing_df2)}")

        # 获取所有列的并集
        all_columns = ["状态"] + list(df1.columns)
        for col in df2.columns:
            if col not in all_columns:
                all_columns.append(col)
        data_columns = all_columns[1:]

        if not key_columns:
            return ExcelDiffer._compare_by_sequence(df1, df2, data_columns, all_columns)
        else:
            return ExcelDiffer._compare_by_keys(df1, df2, key_columns, data_columns, all_columns)

    @staticmethod
    def _compare_by_sequence(df1, df2, data_columns, all_columns):
        """高度矢量化的基于序列对齐的比对逻辑"""
        def get_row_data(df, target_cols):
            rows = []
            for _, row in df.iterrows():
                row_vals = []
                for col in target_cols:
                    val = row.get(col, "")
                    row_vals.append(str(val).strip())
                rows.append(tuple(row_vals))
            return rows

        rows1 = get_row_data(df1, data_columns)
        rows2 = get_row_data(df2, data_columns)
        
        matcher = difflib.SequenceMatcher(None, rows1, rows2, autojunk=False)
        results = []
        
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'equal':
                for k in range(i2 - i1):
                    row_vals = list(rows1[i1 + k])
                    results.append((["一致"] + row_vals, ["一致"] + row_vals, {'status': 'equal', 'diff_cols': []}))
            
            elif tag == 'replace':
                count1 = i2 - i1
                count2 = j2 - j1
                max_count = max(count1, count2)
                for k in range(max_count):
                    has_row1 = k < count1
                    has_row2 = k < count2
                    if has_row1 and has_row2:
                        r1_tuple = rows1[i1 + k]
                        r2_tuple = rows2[j1 + k]
                        
                        if r1_tuple == r2_tuple:
                            vals = list(r1_tuple)
                            results.append((["一致"] + vals, ["一致"] + vals, {'status': 'equal', 'diff_cols': []}))
                        else:
                            l_row = list(r1_tuple)
                            r_row = list(r2_tuple)
                            diff_cols = []
                            for idx, (v1, v2) in enumerate(zip(r1_tuple, r2_tuple)):
                                if v1 != v2:
                                    diff_cols.append(idx + 1)
                            
                            results.append((["修改"] + l_row, ["修改"] + r_row, {'status': 'modified', 'diff_cols': diff_cols}))
                    elif has_row1:
                        row_vals = list(rows1[i1 + k])
                        results.append((["删除"] + row_vals, ["删除"] + ["" for _ in data_columns], {'status': 'deleted', 'diff_cols': []}))
                    elif has_row2:
                        row_vals = list(rows2[j1 + k])
                        results.append((["新增"] + ["" for _ in data_columns], ["新增"] + row_vals, {'status': 'added', 'diff_cols': []}))
            
            elif tag == 'delete':
                for k in range(i1, i2):
                    row_vals = list(rows1[k])
                    results.append((["删除"] + row_vals, ["删除"] + ["" for _ in data_columns], {'status': 'deleted', 'diff_cols': []}))
            
            elif tag == 'insert':
                for k in range(j1, j2):
                    row_vals = list(rows2[k])
                    results.append((["新增"] + ["" for _ in data_columns], ["新增"] + row_vals, {'status': 'added', 'diff_cols': []}))
                    
        return all_columns, results

    @staticmethod
    def _compare_by_keys(df1, df2, key_columns, data_columns, all_columns):
        """高度矢量化的基于主键的比对逻辑"""
        # 1. 预处理：主键列转字符串并去空格
        d1 = df1.copy()
        d2 = df2.copy()
        for col in key_columns:
            d1[col] = d1[col].astype(str).str.strip()
            d2[col] = d2[col].astype(str).str.strip()

        # 2. 使用 merge 进行对齐
        merged = pd.merge(
            d1, d2, 
            on=key_columns, 
            how='outer', 
            suffixes=('_old', '_new'), 
            indicator=True
        )

        # 3. 矢量化计算每列的显示文本和差异标记
        # 初始化左右两列的显示 DataFrame
        left_display = pd.DataFrame(index=merged.index)
        right_display = pd.DataFrame(index=merged.index)
        
        # 记录哪些行有差异（仅针对 'both' 类型）
        both_mask = merged['_merge'] == 'both'
        has_diff_series = pd.Series(False, index=merged.index)

        for col in data_columns:
            col_old = col + '_old' if (col + '_old') in merged.columns else col
            col_new = col + '_new' if (col + '_new') in merged.columns else col
            
            # 统一转为字符串并去空格进行比对
            v1 = merged[col_old].astype(str).str.strip() if col_old in merged.columns else pd.Series("", index=merged.index)
            v2 = merged[col_new].astype(str).str.strip() if col_new in merged.columns else pd.Series("", index=merged.index)
            
            # 只有在 both 且值不同时才标记差异
            diff_mask = (v1 != v2) & both_mask
            has_diff_series |= diff_mask
            
            # 生成显示文本（保持原始数据，不再添加 >>> <<<）
            l_text = v1.copy()
            r_text = v2.copy()
            
            # 对于 left_only，右侧置空；对于 right_only，左侧置空
            l_text.loc[merged['_merge'] == 'right_only'] = ""
            r_text.loc[merged['_merge'] == 'left_only'] = ""
            
            left_display[col] = l_text
            right_display[col] = r_text
            
            # 记录差异列索引（相对于 data_columns）
            # 我们将在 results 的第 3 个元素（tags）中存储更多信息
            # 原有的 tags 结构: ['modified'] 或 ['equal'] 等
            # 新的 tags 结构: {'status': 'modified', 'diff_cols': [idx1, idx2...]}
            # 但为了保持兼容性，我们先保留 results 的基本结构，稍后在 MainWindow 中处理

        # 4. 计算最终状态
        final_status = pd.Series("", index=merged.index)
        final_status.loc[merged['_merge'] == 'left_only'] = "删除"
        final_status.loc[merged['_merge'] == 'right_only'] = "新增"
        final_status.loc[both_mask & has_diff_series] = "修改"
        final_status.loc[both_mask & ~has_diff_series] = "一致"

        # 5. 构建结果列表
        # 结果格式: (left_row, right_row, diff_info)
        # diff_info 现在是一个字典，包含 status 和 diff_cols (发生修改的列索引列表)
        results = []
        status_list = final_status.tolist()
        left_data = left_display.values.tolist()
        right_data = right_display.values.tolist()
        merge_indicators = merged['_merge'].tolist()
        
        for i in range(len(merged)):
            indicator = merge_indicators[i]
            st = status_list[i]
            
            diff_cols = []
            if indicator == 'both' and st == "修改":
                # 找出该行哪些列有差异
                for col_idx, col in enumerate(data_columns):
                    v1 = left_data[i][col_idx]
                    v2 = right_data[i][col_idx]
                    if v1 != v2:
                        diff_cols.append(col_idx + 1) # +1 是因为结果行第一列是"状态"
            
            diff_info = {
                'status': 'deleted' if indicator == 'left_only' else ('added' if indicator == 'right_only' else ('modified' if diff_cols else 'equal')),
                'diff_cols': diff_cols
            }
                
            results.append(([st] + left_data[i], [st] + right_data[i], diff_info))

        return all_columns, results

    @staticmethod
    def _compare_rows(row1, row2, data_columns):
        """对比单行数据并标记差异"""
        l_row = []
        r_row = []
        has_diff = False
        
        for col in data_columns:
            # 兼容处理：如果 row 中缺少某列，则视为空字符串
            v1 = row1.get(col, "")
            v2 = row2.get(col, "")
            
            # 忽略空白字符的差异
            s1 = str(v1).strip()
            s2 = str(v2).strip()
            
            if s1 != s2:
                l_row.append(f">>> {v1} <<<")
                r_row.append(f">>> {v2} <<<")
                has_diff = True
            else:
                l_row.append(str(v1))
                r_row.append(str(v2))
                
        return l_row, r_row, has_diff


    @staticmethod
    def get_text_diff_plain(old_val, new_val):
        """
        计算文本微观差异并返回易读的纯文本格式 (用于 Excel 批注)
        """
        s1 = str(old_val) if pd.notna(old_val) else ""
        s2 = str(new_val) if pd.notna(new_val) else ""
        if s1 == s2: return s1
        
        s = difflib.SequenceMatcher(None, s1, s2)
        result = []
        for tag, i1, i2, j1, j2 in s.get_opcodes():
            if tag == 'equal':
                result.append(s1[i1:i2])
            elif tag == 'delete':
                result.append(f"[-{s1[i1:i2]}-]")
            elif tag == 'insert':
                result.append(f"[+{s2[j1:j2]}+]")
            elif tag == 'replace':
                result.append(f"[-{s1[i1:i2]}-][+{s2[j1:j2]}+]")
        return "".join(result)

    @staticmethod
    def export_diff(output_path, columns, results, key_columns=None):
        """
        将比对结果导出到 Excel，包含带格式的视图和差异清单。
        增强：支持背景高亮、单元格批注、冻结窗格和自动列宽。
        """
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        
        # --- 样式定义 ---
        header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid") # 雅致蓝
        modified_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # 浅橙 (修改)
        added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # 浅绿 (新增)
        deleted_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # 浅红 (删除)
        
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # --- Sheet 1: 对比视图 ---
        ws1 = wb.active
        ws1.title = "对比视图"
        ws1.freeze_panes = "A2" # 冻结首行
        
        half_cols = len(columns)
        # 写入表头
        for i, col in enumerate(columns):
            # 左侧 (旧)
            c1 = ws1.cell(row=1, column=i+1, value=f"旧_{col}")
            c1.fill = header_fill
            c1.font = bold_font
            c1.alignment = center_align
            c1.border = thin_border
            # 右侧 (新)
            c2 = ws1.cell(row=1, column=i+half_cols+2, value=f"新_{col}")
            c2.fill = header_fill
            c2.font = bold_font
            c2.alignment = center_align
            c2.border = thin_border

        # 写入数据
        for row_idx, (left_vals, right_vals, info) in enumerate(results):
            excel_row = row_idx + 2
            status = info.get('status', 'equal')
            diff_cols = info.get('diff_cols', [])
            
            # 左侧
            for col_idx, val in enumerate(left_vals):
                cell = ws1.cell(row=excel_row, column=col_idx+1, value=str(val))
                cell.border = thin_border
                cell.alignment = left_align
                if status == 'deleted':
                    cell.fill = deleted_fill
                elif status == 'modified' and col_idx in diff_cols:
                    cell.fill = modified_fill
                    # 为修改的单元格添加批注显示旧值
                    # cell.comment = Comment(f"原值: {val}", "ExcelDiffer")
            
            # 右侧
            for col_idx, val in enumerate(right_vals):
                cell = ws1.cell(row=excel_row, column=col_idx+half_cols+2, value=str(val))
                cell.border = thin_border
                cell.alignment = left_align
                if status == 'added':
                    cell.fill = added_fill
                elif status == 'modified' and col_idx in diff_cols:
                    cell.fill = modified_fill
                    # 为修改的单元格添加详细差异批注
                    old_val = left_vals[col_idx]
                    diff_text = ExcelDiffer.get_text_diff_plain(old_val, val)
                    cell.comment = Comment(f"差异明细:\n{diff_text}\n\n格式说明: [-删除-] [+新增+]", "ExcelDiffer")

        # 自动调整列宽
        for col_idx in range(1, (half_cols * 2) + 3):
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        # --- Sheet 2: 差异清单 ---
        ws2 = wb.create_sheet("差异清单")
        ws2.freeze_panes = "A2"
        cl_headers = ["主键/位置", "状态", "修改项", "旧值 (旧文件)", "新值 (新文件)", "微观差异说明"]
        for i, h in enumerate(cl_headers):
            cell = ws2.cell(row=1, column=i+1, value=h)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
        
        # 获取主键索引
        key_indices = []
        if key_columns:
            for kc in key_columns:
                if kc in columns:
                    key_indices.append(columns.index(kc))
        
        curr_row = 2
        for i, (left_vals, right_vals, info) in enumerate(results):
            status = info.get('status', 'equal')
            if status == 'equal':
                continue
                
            status_text = "修改" if status == 'modified' else ("新增" if status == 'added' else "删除")
            
            # 行标识
            if key_indices:
                base_vals = right_vals if status != 'deleted' else left_vals
                key_info = " | ".join([str(base_vals[idx]) for idx in key_indices if idx < len(base_vals)])
            else:
                key_info = f"第 {i+1} 行"
            
            if status == 'modified':
                diff_cols = info.get('diff_cols', [])
                for c_idx in diff_cols:
                    col_name = columns[c_idx]
                    l_val = left_vals[c_idx]
                    r_val = right_vals[c_idx]
                    diff_plain = ExcelDiffer.get_text_diff_plain(l_val, r_val)
                    
                    data = [key_info, status_text, col_name, str(l_val), str(r_val), diff_plain]
                    for col_idx, val in enumerate(data):
                        cell = ws2.cell(row=curr_row, column=col_idx+1, value=val)
                        cell.border = thin_border
                        cell.fill = modified_fill
                    curr_row += 1
            else:
                # 新增或删除
                fill = added_fill if status == 'added' else deleted_fill
                data = [key_info, status_text, "整行", "", "", ""]
                if status == 'added':
                    data[4] = "(查看对比视图详情)"
                else:
                    data[3] = "(查看对比视图详情)"
                    
                for col_idx, val in enumerate(data):
                    cell = ws2.cell(row=curr_row, column=col_idx+1, value=val)
                    cell.border = thin_border
                    cell.fill = fill
                curr_row += 1

        for col_idx in range(1, len(cl_headers) + 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

        wb.save(output_path)
