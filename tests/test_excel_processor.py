import unittest
import pandas as pd
import os
import shutil
from src.logic.excel_processor import ExcelDiffer

class TestExcelDiffer(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        """准备测试用的 Excel 文件"""
        cls.test_dir = "test_tmp"
        if not os.path.exists(cls.test_dir):
            os.makedirs(cls.test_dir)
            
        cls.file1 = os.path.join(cls.test_dir, "test1.xlsx")
        cls.file2 = os.path.join(cls.test_dir, "test2.xlsx")
        cls.file_empty = os.path.join(cls.test_dir, "empty.xlsx")
        
        # 创建有偏移表头的测试数据
        df_offset = pd.DataFrame({
            "A": ["ignore", "ignore", "RealCol1", "Data1"],
            "B": ["ignore", "ignore", "RealCol2", "Data2"]
        })
        cls.file_offset = os.path.join(cls.test_dir, "offset.xlsx")
        df_offset.to_excel(cls.file_offset, index=False, header=False)
        
        # 创建标准测试数据
        df1 = pd.DataFrame({
            "ID": [1, 2, 3],
            "Name": ["Alice", "Bob", "Charlie"],
            "Age": [25, 30, 35]
        })
        df1.to_excel(cls.file1, index=False)
        
        # 创建有差异的测试数据 (用于主键比对)
        # 1. 修改 ID=2 的 Age
        # 2. 删除 ID=3
        # 3. 新增 ID=4
        df2 = pd.DataFrame({
            "ID": [1, 2, 4],
            "Name": ["Alice", "Bob", "David"],
            "Age": [25, 31, 40]
        })
        df2.to_excel(cls.file2, index=False)
        
        # 创建空文件 (只有表头)
        df_empty = pd.DataFrame(columns=["A", "B", "C"])
        df_empty.to_excel(cls.file_empty, index=False)

    @classmethod
    def tearDownClass(cls):
        """清理测试文件"""
        if os.path.exists(cls.test_dir):
            shutil.rmtree(cls.test_dir)

    def test_load_sheets(self):
        """测试获取 Sheet 列表"""
        sheets = ExcelDiffer.load_sheets(self.file1)
        self.assertIn("Sheet1", sheets)

    def test_read_excel_raw(self):
        """测试读取 Excel 数据"""
        df = ExcelDiffer.read_excel_raw(self.file1, "Sheet1")
        self.assertEqual(len(df), 3)
        self.assertEqual(list(df.columns), ["ID", "Name", "Age"])

    def test_compare_by_sequence_identical(self):
        """测试序列比对：完全相同"""
        df1 = ExcelDiffer.read_excel_raw(self.file1, "Sheet1")
        all_cols, results = ExcelDiffer.compare_dataframes(df1, df1)
        
        for left, right, info in results:
            self.assertEqual(info['status'], 'equal')
            self.assertEqual(left, right)

    def test_compare_by_sequence_diff(self):
        """测试序列比对：存在差异"""
        df1 = ExcelDiffer.read_excel_raw(self.file1, "Sheet1")
        df2 = ExcelDiffer.read_excel_raw(self.file2, "Sheet1")
        # 序列比对不关注 ID，只按行号对齐
        all_cols, results = ExcelDiffer.compare_dataframes(df1, df2)
        
        # 第三行应该被标记为修改 (Charlie -> David)
        self.assertEqual(results[2][2]['status'], 'modified')

    def test_compare_by_keys(self):
        """测试主键比对"""
        df1 = ExcelDiffer.read_excel_raw(self.file1, "Sheet1")
        df2 = ExcelDiffer.read_excel_raw(self.file2, "Sheet1")
        
        all_cols, results = ExcelDiffer.compare_dataframes(df1, df2, key_columns=["ID"])
        
        # 预期结果：
        # ID=1: equal
        # ID=2: modified (Age: 30 -> 31)
        # ID=3: deleted
        # ID=4: added
        
        status_map = {r[0][1]: r[2]['status'] for r in results if r[0][1] != ""} # r[0][1] 是 ID 列
        # 注意：added 行的 left_row 第一列是空的，ID 在右侧
        added_id = [r[1][1] for r in results if r[2]['status'] == 'added'][0]
        
        self.assertEqual(status_map.get('1'), 'equal')
        self.assertEqual(status_map.get('2'), 'modified')
        self.assertEqual(status_map.get('3'), 'deleted')
        self.assertEqual(str(added_id), '4')

    def test_export_diff_enhanced(self):
        """测试增强型导出功能"""
        df1 = ExcelDiffer.read_excel_raw(self.file1, "Sheet1")
        df2 = ExcelDiffer.read_excel_raw(self.file2, "Sheet1")
        all_cols, results = ExcelDiffer.compare_dataframes(df1, df2, key_columns=["ID"])
        
        export_file = os.path.join(self.test_dir, "export_test.xlsx")
        ExcelDiffer.export_diff(export_file, all_cols, results, key_columns=["ID"])
        
        self.assertTrue(os.path.exists(export_file))
        
        # 验证文件是否可以被 openpyxl 正常打开
        import openpyxl
        wb = openpyxl.load_workbook(export_file)
        self.assertIn("对比视图", wb.sheetnames)
        self.assertIn("差异清单", wb.sheetnames)
        
        ws1 = wb["对比视图"]
        # 验证是否有冻结窗格
        self.assertEqual(ws1.freeze_panes, "A2")
        
    def test_get_text_diff_plain(self):
        """测试纯文本微观差异生成"""
        old_val = "Hello World"
        new_val = "Hello Python"
        plain = ExcelDiffer.get_text_diff_plain(old_val, new_val)
        self.assertIn("[-W-]", plain)
        self.assertIn("[+Pyth+]", plain)
        
    def test_get_text_diff_html(self):
        """测试文本微观差异生成"""
        old_val = "Hello World"
        new_val = "Hello Python"
        html = ExcelDiffer.get_text_diff_html(old_val, new_val)
        
        # 验证是否包含 HTML 标签和关键词片段
        self.assertIn("<span", html)
        self.assertIn("style", html)
        self.assertIn("line-through", html)  # 删除标记
        self.assertIn("Pyth", html)         # 插入的内容片段
        
    def test_read_excel_with_custom_header(self):
        """测试指定表头行号读取"""
        # 指定第 3 行为表头
        df = ExcelDiffer.read_excel_raw(self.file_offset, "Sheet1", header_row=3)
        self.assertEqual(list(df.columns), ["RealCol1", "RealCol2"])
        self.assertEqual(len(df), 1)
        self.assertEqual(df.iloc[0]["RealCol1"], "Data1")

    def test_read_excel_no_header(self):
        """测试无表头模式读取"""
        df = ExcelDiffer.read_excel_raw(self.file1, "Sheet1", has_header=False)
        # file1 有 3 行数据 + 1 行原始表头 = 4 行
        self.assertEqual(len(df), 4)
        self.assertEqual(df.columns[0], "列1")
        # 第一行应该是原始文件的表头行内容
        self.assertEqual(df.iloc[0]["列1"], "ID")

    def test_missing_key_columns(self):
        """测试缺少主键列时的异常处理"""
        df1 = ExcelDiffer.read_excel_raw(self.file1, "Sheet1")
        with self.assertRaises(Exception) as cm:
            ExcelDiffer.compare_dataframes(df1, df1, key_columns=["NonExistent"])
        self.assertIn("缺少关键列", str(cm.exception))

    def test_empty_dataframe(self):
        """测试空数据比对"""
        df_empty = ExcelDiffer.read_excel_raw(self.file_empty, "Sheet1")
        all_cols, results = ExcelDiffer.compare_dataframes(df_empty, df_empty)
        self.assertEqual(len(results), 0)

    def test_compare_all_sheets(self):
        """测试全表自动比对逻辑"""
        # 创建两个多 Sheet 的文件
        file_a = os.path.join(self.test_dir, "multi_a.xlsx")
        file_b = os.path.join(self.test_dir, "multi_b.xlsx")
        
        with pd.ExcelWriter(file_a) as writer:
            pd.DataFrame({"A": [1, 2], "B": [3, 4]}).to_excel(writer, sheet_name="Common", index=False)
            pd.DataFrame({"X": [1]}).to_excel(writer, sheet_name="OnlyA", index=False)
            
        with pd.ExcelWriter(file_b) as writer:
            # Common Sheet 有差异
            pd.DataFrame({"A": [1, 2], "B": [3, 5]}).to_excel(writer, sheet_name="Common", index=False)
            pd.DataFrame({"Y": [1]}).to_excel(writer, sheet_name="OnlyB", index=False)
            
        summary = ExcelDiffer.compare_all_sheets(file_a, file_b)
        
        # 预期：
        # 1. Common: success + stats['modified'] == 1
        # 2. OnlyA: only_in_file1
        # 3. OnlyB: only_in_file2
        
        common_res = [s for s in summary if s['sheet1'] == "Common"][0]
        self.assertEqual(common_res['status'], 'success')
        self.assertEqual(common_res['stats']['modified'], 1)
        
        only_a_res = [s for s in summary if s['sheet1'] == "OnlyA"][0]
        self.assertEqual(only_a_res['status'], 'only_in_file1')
        
        only_b_res = [s for s in summary if s['sheet2'] == "OnlyB"][0]
        self.assertEqual(only_b_res['status'], 'only_in_file2')

if __name__ == '__main__':
    unittest.main()
